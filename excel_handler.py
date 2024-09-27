from config import excel_password
import openpyxl

def add_leads_excel(file_path, leads):
    # Carrega o arquivo Excel
    wb = openpyxl.load_workbook(file_path, keep_links=False)
    
    # Seleciona a planilha 'Plan1'
    sheet = wb['Plan1']

    # Desprotege a planilha se necessário
    if sheet.protection.sheet:
         sheet.protection.set_password(excel_password)
         sheet.protection.sheet = False

    # Procura a primeira linha vazia a partir da linha 2
    start_row = 2
    first_empty_row = None

    for row in range(start_row, sheet.max_row + 2):  # +2 para incluir uma linha a mais caso a última esteja vazia
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, 7)):  # Verifica se todas as colunas da linha estão vazias
            first_empty_row = row
            break

    if first_empty_row is None:
        print("Não há linhas vazias disponíveis.")
        wb.close()
        return

    # Insere os dados a partir da primeira linha vazia encontrada
    for data_row in leads:
        # Preenche as colunas com os valores do dicionário
        sheet.cell(row=first_empty_row, column=1, value=data_row["name"])  # Coluna A
        sheet.cell(row=first_empty_row, column=2, value=data_row["company"])  # Coluna B
        sheet.cell(row=first_empty_row, column=3, value=data_row["campaign"])  # Coluna C
        sheet.cell(row=first_empty_row, column=4, value=data_row["data_invite"])  # Coluna D
        sheet.cell(row=first_empty_row, column=5, value=data_row["linkedin"])  # Coluna E
        sheet.cell(row=first_empty_row, column=6, value=data_row["status"])  # Coluna F
        
        first_empty_row += 1  # Move para a próxima linha vazia

    # Salva as alterações
    wb.save(file_path)
    wb.close()
